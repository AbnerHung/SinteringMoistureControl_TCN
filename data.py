from openpyxl import load_workbook
import torch
import numpy as np


def data_generator():

    input = []

    print('loading data...')
    wb = load_workbook('miniData.xlsx')
    # wb = load_workbook('Data.xlsx')
    sheet = wb["artificial water adding process"]
    for row in sheet.rows:
        data = []
        for cell in row:
            data.append(cell.value)
        data[11] = data[11] + data[12]          # 铁精矿
        data[12] = data[13]                     # 焦炭
        data[13] = data[14]                     # 煤粉
        data[14] = data[15] + data[16]          # 石灰石
        data[15] = data[17] + data[18]          # 白云石
        data[16] = data[19] + data[20] + data[21]  # 返矿
        data[17] = data[22] + data[23]          # 粉尘
        data[18] = data[24] + data[25]          # 生石灰
        input.append(data[1:19])

    print("Loaded successfully !")

    input = input[700:]         # 取700条以后的数据
    input = input[::-1]         # reverse
    '''
    # 测试使用
    for i in range(20):
        print(i, "   ", input[i])
    '''
    input = np.mat(input)
    add_water = input[:, 4:6]
    input = np.delete(input, [4, 5], axis=1)
    
    '''
        input: ['一混测水仪表显示值' '二混测水仪表显示值' '一混测水拟合值' '二混测水拟合值' '1#混匀矿' '2#混匀矿' 
                '3#混匀矿' '4#混匀矿' '#铁精矿铁精矿' '#焦炭' '#煤粉' '#石灰石' '#白云石' '#返矿' '#粉尘' '#生石灰']
        add_water: ['一混加水流量' '二混加水流量']
    '''
    
    return input, add_water


# data_generator()
